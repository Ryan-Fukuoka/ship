import streamlit as st
import openpyxl
from configparser import ConfigParser

def getExcel(joinDf, orderIds, seltNames, sndSec):
	# 讀取設定檔
	mail_service = ConfigParser()
	mail_service.read('mail_service.ini')
	sender = ConfigParser()
	sender.read('sender.ini')
	country = ConfigParser()
	country.read('country.ini')

	# 將訂單資料寫入到郵局出貨單
	rowNum = 3 # 從第3列開始往下增加資料
	wb = openpyxl.load_workbook(filename='formatB.xlsx')
	ws = wb['託運清單']
	for orderId in orderIds:
		orders = joinDf[joinDf['Order ID'] == orderId] # 使用Order ID 找到join 後的table 訂單
		shipName = orders['Ship Name'].unique()[0]
		if shipName not in seltNames:
			continue

		ws['B'+str(rowNum)].value = shipName
		ws['F'+str(rowNum)].value = mail_service['quick']['type']

		# 商城與郵局的國別命名不同，需要在setting.ini 設定對照表
		try:		
			ws['G'+str(rowNum)].value = country['country'][orders['Ship Country'].unique()[0]]
		except KeyError:
			ws['G'+str(rowNum)].value = ''
		ws['H'+str(rowNum)].value = orders['Ship Zipcode'].unique()[0]
		ws['I'+str(rowNum)].value = orders['Ship State'].unique()[0]
		ws['J'+str(rowNum)].value = orders['Ship City'].unique()[0]
		ws['K'+str(rowNum)].value = orders['Ship Address1'].unique()[0] + ' ' + str(orders['Ship Address2'].unique()[0])
		ws['P'+str(rowNum)].value = sender[sndSec]['name']
		ws['R'+str(rowNum)].value = sender[sndSec]['tel']
		ws['S'+str(rowNum)].value = sender[sndSec]['zip']
		ws['T'+str(rowNum)].value = sender[sndSec]['city']
		ws['U'+str(rowNum)].value = sender[sndSec]['address']
		ws['V'+str(rowNum)].value = orders['content'].unique()[0]
		ws['X'+str(rowNum)].value = orders['length'].sum()
		ws['Y'+str(rowNum)].value = orders['width'].sum()
		ws['Z'+str(rowNum)].value = orders['high'].sum()
		ws['AB'+str(rowNum)].value= orders['currency'].unique()[0]

		# 一張訂單多個商品，往右邊移動紀錄
		num = 1
		totalWeight = 0
		for ind, order in orders.iterrows():
			if num == 1:
				ws['AC'+str(rowNum)].value= order['description']
				ws['AD'+str(rowNum)].value= order['Quantity']
				ws['AE'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['AF'+str(rowNum)].value= order['price']
				ws['AG'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 2:
				ws['AJ'+str(rowNum)].value= order['description']
				ws['AK'+str(rowNum)].value= order['Quantity']
				ws['AL'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['AM'+str(rowNum)].value= order['price']
				ws['AN'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 3:
				ws['AQ'+str(rowNum)].value= order['description']
				ws['AR'+str(rowNum)].value= order['Quantity']
				ws['AS'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['AT'+str(rowNum)].value= order['price']
				ws['AU'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 4:
				ws['AX'+str(rowNum)].value= order['description']
				ws['AY'+str(rowNum)].value= order['Quantity']
				ws['AZ'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['BA'+str(rowNum)].value= order['price']
				ws['BB'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 5:
				ws['BE'+str(rowNum)].value= order['description']
				ws['BF'+str(rowNum)].value= order['Quantity']
				ws['BG'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['BH'+str(rowNum)].value= order['price']
				ws['BI'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']	
			elif num == 6:
				ws['BL'+str(rowNum)].value= order['description']
				ws['BM'+str(rowNum)].value= order['Quantity']
				ws['BN'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['BO'+str(rowNum)].value= order['price']
				ws['BP'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 7:
				ws['BS'+str(rowNum)].value= order['description']
				ws['BT'+str(rowNum)].value= order['Quantity']
				ws['BU'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['BV'+str(rowNum)].value= order['price']
				ws['BW'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 8:
				ws['BZ'+str(rowNum)].value= order['description']
				ws['CA'+str(rowNum)].value= order['Quantity']
				ws['CB'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['CC'+str(rowNum)].value= order['price']
				ws['CD'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 9:
				ws['CG'+str(rowNum)].value= order['description']
				ws['CH'+str(rowNum)].value= order['Quantity']
				ws['CI'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['CJ'+str(rowNum)].value= order['price']
				ws['CK'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 10:
				ws['CN'+str(rowNum)].value= order['description']
				ws['CO'+str(rowNum)].value= order['Quantity']
				ws['CP'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['CQ'+str(rowNum)].value= order['price']
				ws['CR'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']

			num += 1

		ws['W'+str(rowNum)].value = totalWeight
		rowNum +=1

	return wb