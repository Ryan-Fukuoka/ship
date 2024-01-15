import streamlit as st
import openpyxl
from configparser import ConfigParser

def getExcel(joinDf, orderIds, seltNames, sndSec):
	# 讀取設定檔
	mail_service = ConfigParser()
	mail_service.read('mail_service.ini')
	sender = ConfigParser()
	sender.read('sender.ini')
	country = ConfigParser()
	country.read('country.ini')

	# 將訂單資料寫入到郵局出貨單
	rowNum = 3 # 從第3列開始往下增加資料
	wb = openpyxl.load_workbook(filename='formatA.xlsx')
	ws = wb['託運清單']
	for orderId in orderIds:
		orders = joinDf[joinDf['Order ID'] == orderId] # 使用Order ID 找到join 後的table 訂單
		shipName = orders['Ship Name'].unique()[0]
		if shipName not in seltNames:
			continue

		ws['B'+str(rowNum)].value = shipName
		ws['F'+str(rowNum)].value = mail_service['normal']['type']

		# 商城與郵局的國別命名不同，需要在setting.ini 設定對照表
		try:		
			ws['G'+str(rowNum)].value = country['country'][orders['Ship Country'].unique()[0]]
		except KeyError:
			ws['G'+str(rowNum)].value = ''
		ws['H'+str(rowNum)].value = orders['Ship Zipcode'].unique()[0]
		ws['I'+str(rowNum)].value = orders['Ship State'].unique()[0]
		ws['J'+str(rowNum)].value = orders['Ship City'].unique()[0]
		ws['K'+str(rowNum)].value = orders['Ship Address1'].unique()[0] + ' ' + str(orders['Ship Address2'].unique()[0])
		ws['P'+str(rowNum)].value = sender[sndSec]['name']
		ws['R'+str(rowNum)].value = sender[sndSec]['tel']
		ws['S'+str(rowNum)].value = sender[sndSec]['zip']
		ws['T'+str(rowNum)].value = sender[sndSec]['city']
		ws['U'+str(rowNum)].value = sender[sndSec]['address']
		ws['V'+str(rowNum)].value = orders['content'].unique()[0]
		ws['X'+str(rowNum)].value = orders['length'].sum()
		ws['Y'+str(rowNum)].value = orders['width'].sum()
		ws['Z'+str(rowNum)].value = orders['high'].sum()
		ws['AA'+str(rowNum)].value= orders['currency'].unique()[0]

		# 一張訂單多個商品，往右邊移動紀錄
		num = 1
		totalWeight = 0
		for ind, order in orders.iterrows():
			if num == 1:
				ws['AB'+str(rowNum)].value= order['description']
				ws['AC'+str(rowNum)].value= order['Quantity']
				ws['AD'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['AE'+str(rowNum)].value= order['price']
				ws['AF'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 2:
				ws['AI'+str(rowNum)].value= order['description']
				ws['AJ'+str(rowNum)].value= order['Quantity']
				ws['AK'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['AL'+str(rowNum)].value= order['price']
				ws['AM'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 3:
				ws['AP'+str(rowNum)].value= order['description']
				ws['AQ'+str(rowNum)].value= order['Quantity']
				ws['AR'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['AS'+str(rowNum)].value= order['price']
				ws['AT'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 4:
				ws['AW'+str(rowNum)].value= order['description']
				ws['AX'+str(rowNum)].value= order['Quantity']
				ws['AY'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['AZ'+str(rowNum)].value= order['price']
				ws['BA'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']
			elif num == 5:
				ws['BD'+str(rowNum)].value= order['description']
				ws['BE'+str(rowNum)].value= order['Quantity']
				ws['BF'+str(rowNum)].value= order['Quantity'] * order['weight']
				ws['BG'+str(rowNum)].value= order['price']
				ws['BH'+str(rowNum)].value= order['Quantity'] * order['price']
				totalWeight += order['Quantity'] * order['weight']	

			num += 1

		ws['W'+str(rowNum)].value = totalWeight
		rowNum +=1

	return wb